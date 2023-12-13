import customtkinter as ctk
import threading as th
import sys
sys.path.append(r'/opt/pylon5')

from pypylon import pylon
import cv2
import xlsxwriter as xl
import os
import pandas as pd
from openpyxl import load_workbook
import numpy as np
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import socket
import json
from time import sleep
import warnings
from PIL import Image, ImageTk

warnings.filterwarnings("ignore")

#Her bliver det defineret hvad ens IPV4 address 
s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
s.bind(('192.168.230.100', 9090))
s.listen(1)

clicked = False
calibrated = False
fig = Figure(figsize=(6, 6))

#Defineret klick funktion globalt mæssigt
def click():
    global clicked
    clicked = True

#definerete funktion med variablerne: (file_path, r, angle) til excel ark
def r_vink_excel(file_path, r, angle):
    global antal
    if os.path.exists(file_path):
        # hvis filen eksister, åben den og load informationen ind i den.
        ex = pd.read_excel(file_path)
        # placer i colums, med antal af colums
        place_column = len(ex.columns)
        antal = 1 + (place_column / 2)
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book

        # Add a new sheet and write your data
        df = pd.DataFrame({f'R: {antal}': r, f'θ: {antal}': angle})
        df.to_excel(writer, sheet_name='Sheet1', index=False, startcol=place_column)

        # gem dokument information()
        writer.close()

    else:
        # hvis filen ikke eksister oprettets der en ny excel fil
        workbook = xl.Workbook(file_path)
        sheet = workbook.add_worksheet()
        antal = 1
        df = pd.DataFrame({f'R: {antal}': r, f'θ: {antal}': angle})
        df.to_excel(file_path, sheet_name='Sheet1', index=False, startcol=0)

#definition af kile posistioner, som gemmes i et seperat excel dokument.
def k_excel(file_path, x):
    if os.path.exists(file_path):
        # If the file exists, open it and load the existing workbook
        ex = pd.read_excel(file_path)
        place_column = len(ex.columns)
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book

        # Add a new sheet and write your data

        df = pd.DataFrame({f'x: {int(place_column + 1)}': x})
        df.to_excel(writer, sheet_name='Sheet1', index=False, startcol=place_column)

        # writer.save()
        writer.close()

    else:
        # If the file doesn't exist, create a new one
        workbook = xl.Workbook(file_path)
        sheet = workbook.add_worksheet()
        antal = 1
        df = pd.DataFrame({f'Position: {antal}': x})
        df.to_excel(file_path, sheet_name='Sheet1', index=False, startcol=0)

#definition af  areal beregninger: hvor der bliver gjordt brug af snørrebåndsformlen
def calculate_area(x, y):
    # Get the number of vertices (assuming x and y have the same length)
    n = len(x)
    # beregniner ved snørrebåndsformlen formlen
    area = 0.5 * abs(sum(x[i] * y[i+1] - x[i+1] * y[i] for i in range(n-1)) + x[n-1] * y[0] - x[0] * y[n-1])
    return area


def calibrate(): #Definering af fisheye kalibreringsfunktion
    global mtx  #Definering af mtx variablen for hele scriptet
    global dist #Definering af dist variablen for hele scriptet
    global calibrated   #Definering af calibrated variablen for hele scriptet
    im1 = cv2.imread(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\works\chess1.jpg")
    im2 = cv2.imread(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\works\chess2.jpg")
    im3 = cv2.imread(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\works\chess3.jpg")
    im4 = cv2.imread(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\works\chess4.jpg")
    im5 = cv2.imread(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\works\chess5.jpg")     #Importer 10 billeder af skakbrætopsætning
    im6 = cv2.imread(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\works\chess6.jpg")
    im7 = cv2.imread(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\works\chess7.jpg")
    im8 = cv2.imread(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\works\chess8.jpg")
    im9 = cv2.imread(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\works\chess9.jpg")
    im10 = cv2.imread(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\works\chess10.jpg")

    img_list = [im1, im2, im3, im4, im5, im6, im7, im8, im9, im10] #Billederne indsættes i en liste
    # criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 30, 0.001)
    objp = np.zeros((6 * 8, 3), np.float32)     #objp bliver defineret som en variabel for en 48x3 matrix
    objp[:, :2] = np.mgrid[0:8, 0:6].T.reshape(-1, 2)   #objp bliver tilpasset et 2D koordinatsystem til at præsentere skakbrættet

    objpoints = []  # 3D point in real world space
    imgpoints = []  # 2D points in image plane

    for image in img_list:  #Dette er et for loop der kalder alle image variabler
        frame = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)  #Denne funktion ændrer farvecoden fra BGR til RGB
        gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)    #Denne funktion danner en gråscalering
        _, binary = cv2.threshold(gray_frame, 100, 255, cv2.THRESH_BINARY)  #Her bliver der tilføjet en binær grænse til det gråscaleret billede
        ret, corners = cv2.findChessboardCorners(binary, (8, 6), None)  #Funktionen finder hjørnerne på billedet og # definerer et (8,6) skakbræt mønster til varibelen corners

        if ret == True: #funktionen tjekker om hjørnerne blev fundet fra sidste funktion
            objpoints.append(objp)  #Varibelen objp bliver vedhæftet til listen objpoints
            imgpoints.append(corners)   #Varibelen corners bliver vedhæftet til listen imgpoints

    # Calibrate the camera
    ret, mtx, dist = cv2.calibrateCamera(objpoints, imgpoints, gray_frame.shape[::-1])    #Funktionen kalibrerer kamaraet ved brug af de tidligere definerede 3D punkter
    print("Calibrated") #Denne funktion printer teksten "Calibrated" når funktionen over er kørt
    calibrated = True   #Variabelen calibrated bliver sat til true når kameraet er kalibraeret
    cv2.destroyAllWindows() #Denne linje lukker alle åbne OpenCV windows efter kalibraringen


def calibrer():     #Definering af funktionen "calibrer"
    if __name__ == "__main__":      #tjekker om sciptet bliver kært som hovedprogram
        p4 = th.Thread(target=calibrate)    #fremstilling af et thread objekt med navnet p4 med funktions målet "calibrate"
        p4.start()      #Starter brugen af threaden


def cam_activate():     #Definering af funktionen "cam_activate"
    if __name__ == "__main__":      #tjekker om sciptet bliver kært som hovedprogram
        p1 = th.Thread(target=camera)       #fremstilling af et thread objekt med navnet p1 med funktions målet "camera"
        p1.start()      #Starter brugen af threaden

#Her bliver x_0 til x_315 defineret til deres neautal position
x_0 = 3.815
x_45 = 3.815
x_90 = 3.815
x_135 = 3.815       
x_180 = 3.815
x_225 = 3.815
x_270 = 3.815
x_315 = 3.815

#defineret matrix_ligning med parameteren: (delta_r_vector)
def matrix_ligning(delta_r_vector):
    global x_0
    global x_90
    global x_180        #De justebare kiler bliver defineret på en måde at alt koden kan se det globalt: (x_0,x_90,x_180,x_270)
    global x_270
    transponeret_Jacobian = np.transpose(Jacobian)      #jacobi matrix

    #np er vores prikprodukt funktion
    gradientf = np.dot(transponeret_Jacobian, delta_r_vector)

    #Hessian er matrix-matrix produkt af den transponeret_Jacobian og Jacobian
    Hessian = np.dot(transponeret_Jacobian, Jacobian)

    #Hvis Hessian matrix er invertibel, calculere den (steps: s_k) ved at bruge newton-raphson optimiserings metode.
    if np.linalg.det(Hessian) != 0:
        inverse_Hessian = np.linalg.inv(Hessian)
        s_k = -alpha * np.dot(inverse_Hessian, gradientf)
        print(f"s_k: {s_k}")
        #Funktion updetere positionerne (x_0, x_90, x_180, x_270) og beregner de korresponderende micro steps
        mic_step0 = round((s_k[0] / 0.004160747858) / (0.45))
        mic_step90 = round((s_k[1] / 0.004160747858) / (0.45))
        mic_step180 = round((s_k[2] / 0.004160747858) / (0.45))
        mic_step270 = round((s_k[3] / 0.004160747858) / (0.45))

        x_0 = x_0 + s_k[0]
        x_90 = x_90 + s_k[1]
        x_180 = x_180 + s_k[2]
        x_270 = x_180 + s_k[3]

        #opdaterede posisioner med microsteps
        mic_steps = [mic_step0, mic_step90, mic_step180, mic_step270]
        pos = [x_0, x_90, x_180, x_270]
        return mic_steps, pos

    else:
        #hvis ikke hassian er invertibel printer den (eksisterer ikke).
        print("Inverse Hessian eksisterer ikke")

#defineret funktion, med parameterne: (theta, r, trin)
def trin_vinkel(theta, r, trin):
    #der bliver lavet denne "vals" som er en liste med værdiger repræcenteret med en vinkel med en givet step
    vals = [h for h in range(0, 360, trin)]
    #For hver værdig in vals, finder den, den vinkel som er tættes på den korresponderne distance i dataen (theta and r)
    theta_ordered = [] #funktionen retunere en liste til vinkel  
    r_ordered = [] #funktionen retunere en liste til radius
    for c in vals: #Et loop 
        min_diff = float('inf')  # Initialize with positive infinity
        theta_or = None #Denne variabel vil blive brugt til at gemme vinkelen med minimeret difference til ønskede vinkel
        r_or = None #ligeledes det samme som vinkel bare med radius

        #Loop der starter over: vinkel(s) og distancen (rad) 
        for s, rad in zip(theta, r):
            diff = abs(c - s) #beregner den absolute forskel c= den ønskede vinkel og s=som er den nuværende vinkel, hvor der kommer ud med en forskel. (note: positive or negative er lige meget for funktionen)
            if diff < min_diff: #linen tjekker for om differancen er mindre en den nuværende mindste diff (min_diff). Hvis ja, så køre koden videre med statmentsene under
                min_diff = diff #updater den minimale differance = diff til den nuværende diff
                theta_or = s #updatere variablen
                r_or = rad #updatere variablen

        theta_ordered.append(theta_or)
        r_ordered.append(r_or)
    return r_ordered, theta_ordered
 #note: r_or og theta_or finde vinkel fra den input dataen som minimere den absolute diff til den ønskede vinkel "c"

#file_path_first = r"C:\Users\Juliu\OneDrive\Skrivebord\Position.xlsx"
#placering_start = [3.815, 3.815, 3.815, 3.815]
#k_excel(file_path=file_path_first, x=placering_start)


#defninieret kamera funktion "det der sker når når der trykkes på knappen på tk.inter brugerfladen"
def camera():
    global r
    global angle
    global theta
    global clicked
    global r_max
    global r_min
    global ani_data
    global antal
    global rotations_data
    global r_exc
    W = 5472
    H = 3648
    count = 0
    # Create an instant camera object with the camera device found first
    camera = pylon.InstantCamera(pylon.TlFactory.GetInstance().CreateFirstDevice())

    # Open the camera
    camera.Open()

    # Set the pixel format to BGR8
    camera.PixelFormat = "BGR8"

    # Start grabbing before the loop
    camera.StartGrabbing()

    while True:
        grabResult = camera.RetrieveResult(5000, pylon.TimeoutHandling_ThrowException)

        if grabResult.GrabSucceeded():
            # Access the image data and convert it to a format suitable for OpenCV
            image = grabResult.Array
            image = cv2.rotate(image, cv2.ROTATE_180) #rotere frame
            resized_image = cv2.resize(image, (546 * 2, 364 * 2)) #skalere frame størrelse
            frame = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            resized_frame = cv2.resize(frame, (546 * 2, 364 * 2))


            #If statement of calibrated er = True: 
            if calibrated == True:
                undistorted_frame = cv2.undistort(frame, mtx, dist, None, mtx) #ikke forvrængning framet med at bruge kamerats parameter
                gray_frame = cv2.cvtColor(undistorted_frame, cv2.COLOR_BGR2GRAY) #conventere ikke forvrængning framet til en grayscale
                blur = cv2.GaussianBlur(gray_frame, (3, 3), 0) #anvender "Gaussian blur" til grayscale billedet til at reducere støj. 
                # blur = cv2.GaussianBlur(gray_frame, (5, 5), 0)
                _, binary = cv2.threshold(blur, 100, 255, cv2.THRESH_BINARY) #Anvender framet binær så alle pixels under 100 bliver 0(sort) og alle over 100 bliver 255(hvide)
                # Display the frame
                contours, hierarchy = cv2.findContours(binary, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE) #finder conturen i den binart image, ved at bruge de to funktioner (binary, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
                contours = list(contours) #conventere contueren til en list
                contours.remove(contours[0]) #Linjen fjerner den første kontur, da det er selve rammen af framet "firekant"
                #ydre_contour = contours[0]



                cv2.drawContours(undistorted_frame, contours, -1, (0, 0, 255), 5) #Linjen laver vores konturen som fremkommer som en rød linje
                cv2.circle(undistorted_frame, (2736, 1824), radius=5, color=(0, 0, 255), thickness=-1) #tegner den fede markerete dot i midten af skærmen
                cv2.line(undistorted_frame, (2900, 1824), (5000, 1824), color=(0, 0, 255), thickness=5) #tegner en red line på det ikke forvrængede frame
                cv2.namedWindow("Undistorted", cv2.WINDOW_NORMAL) # laver en navngivet vindue fane med normale størrelse forhold
                cv2.resizeWindow("Undistorted", 819, 546) # resizer størrelsen af vinduet
                cv2.imshow('Undistorted', undistorted_frame) # displater den endelige frame i det ikke forvrængede frame i et vindue

            #If statement of calibrated er = not true så sker dette  
            else:
                #De meste er kopi pase fra før "pånær nogle variabler"
                #gray_frame = cv2.cvtColor(resized_frame, cv2.COLOR_BGR2GRAY)
                gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY) 
                blur = cv2.GaussianBlur(gray_frame, (3, 3), 0) 
                # blur = cv2.GaussianBlur(gray_frame, (5, 5), 0)
                _, binary = cv2.threshold(blur, 100, 255, cv2.THRESH_BINARY)
                # Display the frame
                contours, hierarchy = cv2.findContours(binary, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
                contours = list(contours)
                contours.remove(contours[0])

                #cv2.drawContours(resized_image, contours, -1, (0, 0, 255), 1)
                cv2.drawContours(image, contours, -1, (0, 0, 255), 5)
                cv2.namedWindow("image", cv2.WINDOW_NORMAL)
                #cv2.resizeWindow("image", 819, 546)
                cv2.imshow("image", image)

            # hvis trykkede = true, så sker dette
            if clicked == True:
                #count = count + 1
                x_list = [] #liste som gemmmer x koordinator
                y_list = [] #liste som gemmmer y koordinator
                max_y = float("-inf")
                min_y = float("inf")
                max_x = float("-inf")
                min_x = float("inf")

                #Loop gennem konturs og samler points
                for contour in contours:
                    for point in contour[:, 0]:
                        x, y = point
                        max_y = max(max_y, y)
                        min_y = min(min_y, y)
                        max_x = max(max_x, x)
                        min_x = min(min_x, x)
                        x_list.append(x)
                        y_list.append(y)

                #kalkulere den distance i y og x retninger (dist)
                dist_y = max_y - min_y
                dist_x = max_x - min_x

                #shift points relative til et referance points (2736, 1824)
                x_list_shifted = []
                y_list_shifted = []
                for a, b in zip(x_list, y_list):
                    x_list_shifted.append(a - (2736))
                    y_list_shifted.append(b - (1824))

                #beregner polar koordinater
                r = np.hypot(x_list_shifted, y_list_shifted) * (1 / 34)
                angle = np.arctan2(y_list_shifted, x_list_shifted)

                theta = [] #tom liste af theta til at gemme
                for vin in angle:
                    omvendt = -1 * vin
                    #posetiv
                    if omvendt > 0:   #omvendt er er større end 0
                        deg = omvendt * (180 / np.pi) 
                        theta.append(deg)
                    #negativ
                    elif omvendt < 0: #omvendt er mindre en nul
                        deg = 360 - (-1 * omvendt * (180 / np.pi))
                        theta.append(deg)
                    #nul
                    elif omvendt == 0:
                        theta.append(0)


                zipped_theta_r = list(zip(theta, r)) #kombinere den konventerede vinkel med den orginale radial distance "r"
                sorted_theta_r = sorted(zipped_theta_r, key=lambda x: x[0]) #listen sortede baseret på vinkel
                theta, r = zip(*sorted_theta_r) # theta, og r er sotede i hver separate liste

                #datapunkter = 40
                #step_data_antal = len(r) // datapunkter
                #r_40_list = [r[i] for i in range(0, len(r), step_data_antal)]

                #
                r_max = float("-inf")
                r_min = float("inf")

                for f, d in zip(r, theta):
                    if f > r_max:
                        r_max = f
                    if f < r_min:
                        r_min = f

                #
                vals_retninger = trin_vinkel(theta=theta, r=r, trin=45)
                r_0 = vals_retninger[0][0]
                r_45 = vals_retninger[0][1]
                r_90 = vals_retninger[0][2]
                r_135 = vals_retninger[0][3]
                r_180 = vals_retninger[0][4]
                r_225 = vals_retninger[0][5]
                r_270 = vals_retninger[0][6]
                r_315 = vals_retninger[0][7]

                omkreds = np.sum(np.sqrt(np.diff(x_list) ** 2 + np.diff(y_list) ** 2))

                # Numerical integration for area using the trapezoidal rule
                area = calculate_area(x_list, y_list)

                Rundhed = (4 * np.pi * area) / (omkreds ** 2) * 100

                labelr0.configure(text=f"{round(r_0, 4)}")
                labelr45.configure(text=f"{round(r_45, 4)}")
                labelr90.configure(text=f"{round(r_90, 4)}")
                labelr135.configure(text=f"{round(r_135, 4)}")
                labelr180.configure(text=f"{round(r_180, 4)}")
                labelr225.configure(text=f"{round(r_225, 4)}")
                labelr270.configure(text=f"{round(r_270, 4)}")
                labelr315.configure(text=f"{round(r_315, 4)}")

                labelRundhed.configure(text=f"Rundhed: {round(Rundhed, 3)}%")

                ani_data = [r_0, r_45, r_90, r_135, r_180, r_225, r_270, r_315]

                file_path = r"C:\Users\Juliu\OneDrive\Skrivebord\Data_rad_vink.xlsx"
                file_path1 = r"C:\Users\Juliu\OneDrive\Skrivebord\Position.xlsx"
                exce = trin_vinkel(theta=theta, r=r, trin=1)
                r_exc = exce[0]
                theta_exc = exce[1]

                delta_r = []
                for i, a in zip(r_exc, r_ref):
                    delta_r.append(i - a)

                rotations_data_matrix = matrix_ligning(delta_r_vector=delta_r)
                rotations_data = rotations_data_matrix[0]
                posi = rotations_data_matrix[1]
                print(f"Micro steps: {rotations_data}")


                r_vink_excel(file_path=file_path, r=r_exc, angle=theta_exc)
                k_excel(file_path=file_path1, x=posi)

                move_canv()

                clicked = False



            elif cv2.waitKey(1) == ord("q"):
                break

            grabResult.Release()

    # Stop grabbing and close the camera and OpenCV window
    camera.StopGrabbing()
    camera.Close()
    cv2.destroyAllWindows()


def polar_plot():
    fig = plt.figure(figsize=(6, 6), facecolor="dimgray")
    ax = plt.subplot(111, projection='polar', facecolor="dimgray")
    ax.clear()
    if antal == 1:
        ax.set_title(f"Emne: 1", font="Times New Roman", fontsize=14)

    elif antal > 1:
        ax.set_title(f"Emne: {int(antal)}", font="Times New Roman", fontsize=14)

    ax.set_rlim((int(r_min) - 2), (int(r_max) + 2))
    #ax.set_theta_direction(-1)
    ax.xaxis.grid(True, color="white")
    ax.yaxis.grid(True, color="white")
    ax.yaxis.set_tick_params(labelcolor="white")

    custom_labels = ['0°', '45°', '90°', '135°', '180°', '225°', '270°', '315°']
    #custom_labels = ['0°', '315°', '270°', '225°', '180°', '135°', '90°', '45°']
    ax.set_thetagrids(range(0, 360, 45), labels=custom_labels, color="white")

    ax.plot(np.radians(theta), r, color="red")
    ax.plot(np.radians(theta_ref), r_ref, color="black")
    canvas = FigureCanvasTkAgg(fig, master=frame1)
    canvas.get_tk_widget().grid(row=1, column=1, pady=5, padx=5)


def send_data():
    data = rotations_data
    data_dump = json.dumps(data)
    data_size = len(data_dump)
    clientsocket.send("data size".encode('utf-8'))
    clientsocket.send(str(data_size).encode('utf-8'))
    sleep(1)
    clientsocket.send(data_dump.encode('utf-8'))


def justering():
    clientsocket.send("justering".encode('utf-8'))


def con():
    global clientsocket

    while True:
        clientsocket, address = s.accept()


def draw_oval(canvas, midpoint, radius, color):
    x, y = midpoint
    left = x - radius
    top = y - radius
    right = x + radius
    bottom = y + radius
    canvas.create_oval(left, top, right, bottom, fill=color)

 #
def move_canv():
    dx0 = -1 * round((x_0 - 3.815) * (480 / 262))
    dy90 = round((x_90 - 3.815) * (480 / 262))
    dx180 = round((x_180 - 3.815) * (480 / 262))
    dy270 = -1 * round((x_270 - 3.815) * (480 / 262))

    canvas.move(kile0, dx0, 0)
    canvas.move(lin0, dx0, 0)

    canvas.move(kile90, 0, dy90)
    canvas.move(lin90, 0, dy90)

    canvas.move(kile180, dx180, 0)
    canvas.move(lin180, dx180, 0)

    canvas.move(kile270, 0, dy270)
    canvas.move(lin270, 0, dy270)

#Denne kode definere
def indstil_canv():
    #canvas.move(kile0, -7, 0)
    #canvas.move(lin0, -7, 0)
    #canvas.move(kile90, 0, 7)
    #canvas.move(lin90, 0, 7)
    #canvas.move(kile180, 7, 0)
    #canvas.move(lin180, 7, 0)
    #canvas.move(kile270, 0, -7)
    #canvas.move(lin270, 0, -7)

    #Denne kode skaber et animationsloop som bevæger variblerne Klie0 og lin0 i små steps
    n = 0
    while n < 7:
        n += 1
        canvas.move(kile0, -1, 0)
        canvas.move(lin0, -1, 0)
        sleep(0.2)  #Denne linje skaber et delay på 0.2 sekunder efter hver bevægelse

    # Denne kode skaber et animationsloop som bevæger variblerne Klie90 og lin90 i små steps
    n = 0
    while n < 7:
        n += 1
        canvas.move(kile90, 0, 1)
        canvas.move(lin90, 0, 1)
        sleep(0.2)  #Denne linje skaber et delay på 0.2 sekunder efter hver bevægelse

    # Denne kode skaber et animationsloop som bevæger variblerne Klie180 og lin180 i små steps
    n = 0
    while n < 7:
        n += 1
        canvas.move(kile180, 1, 0)
        canvas.move(lin180, 1, 0)
        sleep(0.2)  #Denne linje skaber et delay på 0.2 sekunder efter hver bevægelse


    # Denne kode skaber et animationsloop som bevæger variblerne Klie270 og lin270 i små steps
    n = 0
    while n < 7:
        n += 1
        canvas.move(kile270, 0, -1)
        canvas.move(lin270, 0, -1)
        sleep(0.2)  #Denne linje skaber et delay på 0.2 sekunder efter hver bevægelse

#Koden under bliver brugt til at skaber en brugergrænseoverflade, defineret under navnet GUI
def gui():
    #En masse varibeler bliver globalt defineret
    global frame1
    global count
    global labelr0
    global labelr45
    global labelr90
    global labelr135
    global labelr180
    global labelr225
    global labelr270
    global labelr315
    global labelRundhed
    global canvas
    global kile0
    global kile90
    global kile180
    global kile270
    global lin0
    global lin90
    global lin180
    global lin270

    #BrugerGrænseOverfladen bliver sat til dark mode
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("dark-blue")

    #BrugerGrænseOverfladen bliver tildelt en pixel størrelse på (1230x550)
    window = ctk.CTk()
    window.geometry("1230x550")

    #Der bliver defineret et icon som bliver resizet til en spcifik størrelse.
    cam = Image.open(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\icons\camera-icon.png")
    cam = cam.resize((28, 28), Image.ANTIALIAS)
    tk_cam = ImageTk.PhotoImage(cam)

    # Der bliver defineret et icon som bliver resizet til en spcifik størrelse.
    stepper = Image.open(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\icons\stepper.png")
    stepper = stepper.resize((20, 20), Image.ANTIALIAS)
    tk_stepper = ImageTk.PhotoImage(stepper)

    # Der bliver defineret et icon som bliver resizet til en spcifik størrelse.
    send = Image.open(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\icons\send.png")
    send = send.resize((20, 20), Image.ANTIALIAS)
    tk_send = ImageTk.PhotoImage(send)

    # Der bliver defineret et icon som bliver resizet til en spcifik størrelse.
    cali = Image.open(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\icons\calibrate.png")
    cali = cali.resize((20, 20), Image.ANTIALIAS)
    tk_cali = ImageTk.PhotoImage(cali)

    # Der bliver defineret et icon som bliver resizet til en spcifik størrelse.
    excl = Image.open(r"C:\Users\Juliu\OneDrive\Skrivebord\MP3\icons\excel.png")
    excl = excl.resize((20, 20), Image.ANTIALIAS)
    tk_excl = ImageTk.PhotoImage(excl)
    #count = ctk.IntVar()
    #count.set(0)

    #Dette stykke kode bliver brugt til at skabe 6 frames til at oganisere layoutet i BrugerGrænseOverfladen.
    #frame er til knapperne
    frame = ctk.CTkFrame(master=window, width=300, height=300)
    frame.grid(row=0, column=0, pady=5, padx=5)

    #frame1 er til det polære plot
    frame1 = ctk.CTkFrame(master=window, width=500, height=500)
    frame1.grid(row=0, column=1, pady=5, padx=5)

    #frame2 er til vinkel kolonnen
    frame2 = ctk.CTkFrame(master=window, width=250, height=480)
    frame2.grid(row=0, column=2, pady=5, padx=5)

    #frame3 er til Radius kolonnen
    frame3 = ctk.CTkFrame(master=window, width=250, height=480)
    frame3.grid(row=0, column=3, pady=5, padx=5)

    #frame4 er til runhedssøjlen under det polære plot
    frame4 = ctk.CTkFrame(master=window)
    frame4.grid(row=1, column=1)

    #frame5 er til værktøjets projektion
    frame5 = ctk.CTkFrame(master=window, width=500, height=500)
    frame5.grid(row=0, column=4)

    #Her bliver der skabt en knap funktion som gemmer data, som gør brug af iconnet gemt fra tidligere
    button = ctk.CTkButton(master=frame,
                           text="Gem Data",
                           font=("Times New Roman", 15),
                           fg_color="dimgray",
                           image=tk_excl,
                           command=click)
    button.grid(row=0, column=0, pady=5,padx=5)  # Denne linje beskriver hvilken række og kolonne knappen og teksten bliver sat.

    # Her bliver der skabt en knap funktion som skaber et polært plot, som gør brug af iconnet gemt fra tidligere
    button1 = ctk.CTkButton(master=frame,
                            text="Polar Plot",
                            font=("Times New Roman", 15),
                            command=polar_plot,
                            fg_color="dimgray")
    button1.grid(row=1, column=0, pady=5, padx=5)   #Denne linje beskriver hvilken række og kolonne knappen og teksten bliver sat.

    # Her bliver der skabt en knap funktion som sender micro-steps beregningerne til SBC'en,
    # som gør brug af iconnet gemt fra tidligere
    button2 = ctk.CTkButton(master=frame,
                            font=("Times New Roman", 15),
                            text="Send",
                            image=tk_send,
                            fg_color="dimgray",
                            command=send_data)
    button2.grid(row=2, column=0, pady=5, padx=5)   #Denne linje beskriver hvilken række og kolonne knappen og teksten bliver sat.

    # Her bliver der skabt en knap funktion som sender kommandoen med at justere de microsteps som er beregnet på motoren,
    # som gør brug af iconnet gemt fra tidligere
    button3 = ctk.CTkButton(master=frame,
                            font=("Times New Roman", 15),
                            text="Justere",
                            image=tk_stepper,
                            fg_color="dimgray",
                            command=justering)
    button3.grid(row=3, column=0, pady=5, padx=5)   #Denne linje beskriver hvilken række og kolonne knappen og teksten bliver sat.

    # Her bliver der skabt en knap funktion som bliver brugt til at kalibrere kameraet,
    # som gør brug af iconnet gemt fra tidligere
    button4 = ctk.CTkButton(master=frame,
                            font=("Times New Roman", 15),
                            text="Calibrer",
                            image=tk_cali,
                            fg_color="dimgray",
                            command=calibrer)
    button4.grid(row=4, column=0, pady=5, padx=5)   #Denne linje beskriver hvilken række og kolonne knappen og teksten bliver sat.

    # Her bliver der skabt en knap funktion som aktivere kameraet, som gør brug af iconnet gemt fra tidligere
    button5 = ctk.CTkButton(master=frame,
                            font=("Times New Roman", 15),
                            text="Camera",
                            fg_color="dimgray",
                            image=tk_cam,
                            command=cam_activate)
    button5.grid(row=5, column=0, pady=5, padx=5)   #Denne linje beskriver hvilken række og kolonne knappen og teksten bliver sat.

    #Denne kode bliver brugt til at skabe tekst der skal stå på en kolonne
    labelv = ctk.CTkLabel(master=frame2, text="Vinkel")
    label0 = ctk.CTkLabel(master=frame2, text='0°')
    label45 = ctk.CTkLabel(master=frame2, text='45°')
    label90 = ctk.CTkLabel(master=frame2, text='90°')
    label135 = ctk.CTkLabel(master=frame2, text='135°')
    label180 = ctk.CTkLabel(master=frame2, text='180°')
    label225 = ctk.CTkLabel(master=frame2, text='225°')
    label270 = ctk.CTkLabel(master=frame2, text='270°')
    label315 = ctk.CTkLabel(master=frame2, text='315°')

    #Denne kode bliver brugt til at indsætte tekst labelerne skabt over ind på forskellige rækker
    labelv.grid(row=0, column=0, pady=10, padx=10)
    label0.grid(row=1, column=0, pady=10, padx=10)
    label45.grid(row=2, column=0, pady=10, padx=10)
    label90.grid(row=3, column=0, pady=10, padx=10)
    label135.grid(row=4, column=0, pady=10, padx=10)
    label180.grid(row=5, column=0, pady=10, padx=10)
    label225.grid(row=6, column=0, pady=10, padx=10)
    label225.grid(row=7, column=0, pady=10, padx=10)
    label270.grid(row=8, column=0, pady=10, padx=10)
    label315.grid(row=9, column=0, pady=10, padx=10)

    # Denne kode bliver brugt til at lable radiusen fra forskellige vinker der skal stå på en kolonne
    labelr = ctk.CTkLabel(master=frame3, text="Radius")
    labelr0 = ctk.CTkLabel(master=frame3, text=f"")
    labelr45 = ctk.CTkLabel(master=frame3, text=f"")
    labelr90 = ctk.CTkLabel(master=frame3, text=f"")
    labelr135 = ctk.CTkLabel(master=frame3, text=f"")
    labelr180 = ctk.CTkLabel(master=frame3, text=f"")
    labelr225 = ctk.CTkLabel(master=frame3, text=f'')
    labelr270 = ctk.CTkLabel(master=frame3, text=f'')
    labelr315 = ctk.CTkLabel(master=frame3, text=f'')

    # Denne kode indsætter radius værdierne på rækkerne i en kolonne
    labelr.grid(row=0, column=0, pady=10, padx=10)
    labelr0.grid(row=1, column=0, pady=10, padx=10)
    labelr45.grid(row=2, column=0, pady=10, padx=10)
    labelr90.grid(row=3, column=0, pady=10, padx=10)
    labelr135.grid(row=4, column=0, pady=10, padx=10)
    labelr180.grid(row=5, column=0, pady=10, padx=10)
    labelr225.grid(row=6, column=0, pady=10, padx=10)
    labelr225.grid(row=7, column=0, pady=10, padx=10)
    labelr270.grid(row=8, column=0, pady=10, padx=10)
    labelr315.grid(row=9, column=0, pady=10, padx=10)

    #Denne kode bruges til at skabe en tabel til rundhed hvori en beregnet rundhed bliver sat i
    labelRundhed = ctk.CTkLabel(master=frame4, text=f"Rundhed:")
    labelRundhed.grid(row=0, column=0, pady=10, padx=10)

    #Denne kode definere varibelen canvas til at være frame5 og giver den dimensionerne 500x500
    canvas = ctk.CTkCanvas(frame5, width=500, height=500, bg='grey25')
    canvas.pack(padx=10, pady=10)


    #Koden under tegner en oval på et lærred, Ovalen beskriver værktøjet
    draw_oval(canvas=canvas, midpoint=(250, 250), radius=240, color="grey70")
    draw_oval(canvas=canvas, midpoint=(250, 250), radius=194, color="grey38")
    canvas.create_rectangle(227, 42, 273, 458, fill="grey70")
    canvas.create_rectangle(42, 227, 458, 273, fill="grey70")
    draw_oval(canvas=canvas, midpoint=(250, 250), radius=68, color="grey25")
    #Koden under skaber 4 rectangler på placeringerne 0°,90°,180°,270° indenfor en oval.
    # disse rectangler bliver defineret som kiler, og bliver brugt til at vise kilens placering i værktøjet.
    kile0 = canvas.create_rectangle(337, 227, 458, 273, fill="grey", tags='draggable')
    position0 = (470, 250)
    canvas.create_text(position0, text="0°", font=("Times New Roman", 12))
    kile90 = canvas.create_rectangle(227, 42, 273, 163, fill="grey")
    position90 = (250, 30)
    canvas.create_text(position90, text="90°", font=("Times New Roman", 12))
    kile180 = canvas.create_rectangle(42, 227, 163, 273, fill="grey")
    position180 = (26, 250)
    canvas.create_text(position180, text="180°", font=("Times New Roman", 12))
    kile270 = canvas.create_rectangle(227, 337, 273, 458, fill="grey")
    position270 = (250, 470)
    canvas.create_text(position270, text="270°", font=("Times New Roman", 12))

    #Denne kode skaber en stiplede linje på lærredet fra start koordinaterne til slut koordinaterne
    start0 = (391, 220)
    end0 = (391, 280)
    line0 = canvas.create_line(start0, end0, dash=(5, 2))

    #Denne kode skaber en stiplede linje på lærredet fra start koordinaterne til slut koordinaterne
    start90 = (220, 109)
    end90 = (280, 109)
    line90 = canvas.create_line(start90, end90, dash=(5, 2))

    #Denne kode skaber en stiplede linje på lærredet fra start koordinaterne til slut koordinaterne
    start180 = (109, 220)
    end180 = (109, 280)
    line180 = canvas.create_line(start180, end180, dash=(5, 2))

    #Denne kode skaber en stiplede linje på lærredet fra start koordinaterne til slut koordinaterne
    start270 = (220, 391)
    end270 = (280, 391)
    line270 = canvas.create_line(start270, end270, dash=(5, 2))

    #Denne kode skaber en cyan farvet stiplede linje på lærredet fra start koordinaterne til slut koordinaterne.
    #Koden beskriver en linje der beskriver midten af kilen,-
    # hvilket bliver brugt til at vise hvorlangt kilen er fra dens neutral akse
    start0 = (398, 220)
    end0 = (398, 280)
    lin0 = canvas.create_line(start0, end0, dash=(5, 2), fill="cyan2")

    # Denne kode skaber en cyan farvet stiplede linje på lærredet fra start koordinaterne til slut koordinaterne.
    # Koden beskriver en linje der beskriver midten af kilen,-
    # hvilket bliver brugt til at vise hvorlangt kilen er fra dens neutral akse
    start90 = (220, 102)
    end90 = (280, 102)
    lin90 = canvas.create_line(start90, end90, dash=(5, 2), fill="cyan2")

    # Denne kode skaber en cyan farvet stiplede linje på lærredet fra start koordinaterne til slut koordinaterne.
    # Koden beskriver en linje der beskriver midten af kilen,-
    # hvilket bliver brugt til at vise hvorlangt kilen er fra dens neutral akse
    start180 = (102, 220)
    end180 = (102, 280)
    lin180 = canvas.create_line(start180, end180, dash=(5, 2), fill="cyan2")

    # Denne kode skaber en cyan farvet stiplede linje på lærredet fra start koordinaterne til slut koordinaterne.
    # Koden beskriver en linje der beskriver midten af kilen,-
    # hvilket bliver brugt til at vise hvorlangt kilen er fra dens neutral akse
    start270 = (220, 398)
    end270 = (280, 398)
    lin270 = canvas.create_line(start270, end270, dash=(5, 2), fill="cyan2")
    # Bind mouse events
    #canvas.tag_bind('draggable', '<ButtonPress-1>', on_press)
    #canvas.tag_bind('draggable', '<B1-Motion>', on_drag)

    #Denne kode tilader GUI til at give en respons til bruger-interaktion
    window.mainloop()

#Koden under skaber threads, som bliver kørt samtidigt
if __name__ == "__main__":      #Der bliver tjekket om scriptet bliver kørt som hovedprogram
    p2 = th.Thread(target=gui)      #Der bliver skabt en ny thread kaldt (p2) der tildeler "target" variebelen til gui funktionen
    p3 = th.Thread(target=con)      #Der bliver skabt en ny thread kaldt (p3) der tildeler "target" variebelen til con funktionen

    p2.start()      #Dette starter p2 threaden, som blev skabt i linjerne over
    p3.start()      #Dette starter p3 threaden, som blev skabt i linjerne over

    sleep(3)        #Denne kommando pauser main threaden i 3 sekunder, hvilket tilader at gui og con threadende kan starte med at initialisere
    indstil_canv()
